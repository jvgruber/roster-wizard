import calendar
import re
from collections import defaultdict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os


def extractInformation(answerset):
    """
    Extracts the name and day information from an answer set string.

    Args:
        answerset (str): The input answer set string containing assignments in the format "assigned(name,day)".

    Returns:
        dict: A dictionary where keys are names (capitalized) and values are lists of days (e.g., {'Matthias': ['d1', 'd2']}).
    """
    results = re.findall(r'\((.*?)\)', answerset)

    date_info = defaultdict(list)

    for result in results:
        name, value = result.split(',')
        name = name.capitalize()  # Capitalize the name
        date_info[name].append(value)

    date_info = dict(date_info)

    return date_info


def getYear():
    """
    Gets the current year.

    Returns:
        int: The current year.
    """
    return date.today().year


def days_in_month(month):
    """
    Calculates the number of days in a given month for the current year.

    Args:
        month (int): The month as an integer (1-12).

    Returns:
        tuple: A tuple containing the number of days in the month and the current year.
    """
    year = getYear()
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return (next_month - date(year, month, 1)).days, year


def mapAnswersetToDate(date_info, month):
    """
    Maps days in the answer set (e.g., 'd1') to actual calendar dates for a given month.

    Args:
        date_info (dict): A dictionary with names as keys and list of days (e.g., {'Matthias': ['d1', 'd2']}).
        month (int): The month number (1-12).

    Returns:
        dict: A dictionary with names as keys and lists of actual dates (e.g., {'Matthias': ['2024-12-01']}).
    """
    num_days, year = days_in_month(month)

    for name in date_info:
        mapped_dates = []
        for day_code in date_info[name]:
            if day_code.startswith('d'):
                day = int(day_code[1:])

                if 1 <= day <= num_days:
                    mapped_dates.append(f"{year}-{month:02d}-{day:02d}")
                else:
                    mapped_dates.append("Invalid Date")
        date_info[name] = mapped_dates

    return dict(date_info)


def writeCalendarToExcel(data_dict, month, file_name, directory=None):
    """
    Writes a calendar to an Excel file, including names for each day.

    Args:
        data_dict (dict): A dictionary where keys are names and values are lists of actual dates.
                          Example: {'Matthias': ['2024-12-01', '2024-12-02']}.
        month (int): The month number (1-12).
        file_name (str): The name of the output Excel file.
        directory (str, optional): The directory path where the file should be saved. Defaults to the current directory.

    Returns:
        None
    """
    year = getYear()

    # Determine the full file path
    if directory:
        os.makedirs(directory, exist_ok=True)  # Create directory if it doesn't exist
        full_path = os.path.join(directory, file_name)
    else:
        full_path = file_name

    # Create a new workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"

    # Add month and year as the title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws["A1"] = f"{calendar.month_name[month]} {year}"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Days of the week header
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for col, day in enumerate(days, start=1):
        ws.cell(row=2, column=col).value = day
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")

    # Get the month's calendar matrix
    cal = calendar.monthcalendar(year, month)

    # Prepare a mapping of days to all names
    day_to_names = defaultdict(list)
    for name, dates in data_dict.items():
        for date in dates:
            if date.startswith(f"{year}-{month:02d}"):
                day = int(date.split("-")[2])
                day_to_names[day].append(name)

    # Write the calendar dates and names into the Excel sheet
    row_start = 3
    for week in cal:
        for col, day in enumerate(week, start=1):
            if day != 0:  # Valid day
                cell = ws.cell(row=row_start, column=col)
                cell.value = day

                # Add names if available for this day
                if day in day_to_names:
                    names = "\n".join(day_to_names[day])
                    cell.value = f"{day}\n{names}"
                    cell.alignment = Alignment(wrap_text=True, horizontal="center")
        row_start += 1

    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Save the Excel file
    wb.save(full_path)
    print(f"Calendar saved to {full_path}")


def asp2excel(answerset, month, file_name, directory=None):
    """
    Processes an ASP answer set and writes the data into an Excel calendar.

    Args:
        answerset (str): The input answer set string containing assignments in the format "assigned(name,day)".
        month (int): The month number (1-12).
        file_name (str): The name of the output Excel file.
        directory (str, optional): The directory path where the file should be saved. Defaults to the current directory.

    Returns:
        None
    """
    date_info = extractInformation(answerset)

    mapped_info = mapAnswersetToDate(date_info, month)

    writeCalendarToExcel(mapped_info, month, file_name, directory)
